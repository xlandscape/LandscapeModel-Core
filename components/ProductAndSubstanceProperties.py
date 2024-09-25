"""Class definition for the Landscape Model ProductAndSubstanceProperties component."""
import typing
import openpyxl
import base
import attrib


class ProductAndSubstanceProperties(base.Component):
    def __init__(self, name: str, default_observer: base.Observer, default_store: typing.Optional[base.Store]) -> None:
        super(ProductAndSubstanceProperties, self).__init__(name, default_observer, default_store)
        self._inputs = base.InputContainer(
            self, (base.Input("FilePath", (attrib.Class(str), attrib.Unit(None), attrib.Scales("global"))),))
        self._outputs = base.ProvisionalOutputs(self, default_store)

    def run(self) -> None:
        with open(self.inputs["FilePath"].read().values, "rb") as f:
            workbook = openpyxl.load_workbook(f, True, False, True, False)
            assert [x.value for x in workbook["description"][1]] == ["Property name", "Property value", "Comment"]
            assert [x.value for x in workbook["columns"][1]] == ["Column name", "Data type", "Unit", "Description"]
            description = {
                x[0]: x[1] for x in workbook["description"].iter_rows(max_col=2, min_row=2, values_only=True)}
            assert description["Schema version"] == "0.1"
            assert description["Scale"] in ("chemical/substance", "chemical/product")
            columns = {
                x[0]: x[1:4] for x in workbook["columns"].iter_rows(min_row=2, values_only=True)}
            assert "ElementName" in columns
            assert columns["ElementName"][:2] == ("str", None)
            assert all([x[0] in ("float", "str") for x in columns.values()])
            property_names = [x.value for x in workbook["properties"][1]]
            assert set(property_names) == set(columns.keys())
            properties = {x: [] for x in property_names}
            for j, row in enumerate(workbook["properties"].iter_rows(min_row=2)):
                for i, cell in enumerate(row):
                    value = "" if isinstance(cell, openpyxl.cell.read_only.EmptyCell) else cell.value
                    property_type = columns[property_names[i]][0]
                    if property_type == "str":
                        properties[property_names[i]].append(str(value))
                    elif property_type == "float":
                        if value:
                            properties[property_names[i]].append(float(value))
                        else:
                            self.default_observer.write_message(
                                2,
                                "Missing property value",
                                f'Chemical {properties["ElementName"][j]}, '
                                f'{property_names[i]}'
                            )
                            properties[property_names[i]].append(float("nan"))
                    else:
                        raise ValueError(property_type)
        for column_name, column_description in columns.items():
            self.outputs.append(
                base.Output(
                    column_name,
                    self.default_store,
                    self,
                    {"scales": description["Scale"], "unit": column_description[1]},
                    column_description[2]
                )
            )
            self.outputs[column_name].set_values(properties[column_name], element_names=[self.outputs["ElementName"]])
